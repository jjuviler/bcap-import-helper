#!/usr/bin/env python3
"""
BCAP Import Helper

Consolidates Marketing, Sales, Service, and Website tabs from an XLSX file
into a single Import-Ready tab and exports it as CSV.

Exit codes (CLI only):
  0  — Success
  10 — Missing required tabs (re-run with --skip-missing-tabs to proceed)
  20 — Heading mismatch across tabs (fix the XLSX and re-run)
  1  — General error
"""

import argparse
import csv
import os
import sys
from datetime import datetime, date


try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


REQUIRED_TABS = ["Marketing", "Sales", "Service", "Website"]
FOCUS_KW_NAMES = {"focus keyword", "focus kw"}
ASSIGNMENT_SPRINT = "assignment sprint"

DATE_HEADING_KEYWORDS = {
    "date", "sprint", "due", "timing", "published", "scheduled",
    "launch", "go live", "golive", "go-live",
}

# Columns that must have a value in every active row (rows with a Focus KW value)
REQUIRED_ACTIVE_COLUMNS = [
    "blog team", "action", "topic cluster", "primary keyword",
    "url", "content brief", "timing", "assignee", "content source",
    "property", "assignment sprint", "due date",
]


# ---------------------------------------------------------------------------
# Custom exceptions (used by web app; CLI catches and converts to sys.exit)
# ---------------------------------------------------------------------------

class MissingTabsError(Exception):
    pass

class HeadingMismatchError(Exception):
    pass

class MissingColumnsError(Exception):
    pass


# ---------------------------------------------------------------------------
# Tab discovery
# ---------------------------------------------------------------------------

def find_tabs(wb):
    """Return (found: dict canonical→actual, missing: list)."""
    sheet_map = {s.lower(): s for s in wb.sheetnames}
    found, missing = {}, []
    for tab in REQUIRED_TABS:
        if tab.lower() in sheet_map:
            found[tab] = sheet_map[tab.lower()]
        else:
            missing.append(tab)
    return found, missing


# ---------------------------------------------------------------------------
# Reading sheet data into plain Python lists
# ---------------------------------------------------------------------------

def read_sheet(ws):
    """
    Return all rows as a list of lists (values only).
    Trailing all-None rows are stripped.
    """
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    while rows and all(v is None for v in rows[-1]):
        rows.pop()
    return rows


# ---------------------------------------------------------------------------
# Heading validation
# ---------------------------------------------------------------------------

def validate_headings(found_tabs, wb):
    """
    Ensure row 1 is identical across all present tabs.
    Returns (ok: bool, detail: str).
    """
    tab_headings = {}
    for canonical, actual in found_tabs.items():
        data = read_sheet(wb[actual])
        tab_headings[canonical] = data[0] if data else []

    names = list(tab_headings.keys())
    if len(names) < 2:
        return True, ""

    ref_name = names[0]
    ref = tab_headings[ref_name]
    mismatches = []

    for name in names[1:]:
        other = tab_headings[name]
        if ref != other:
            max_len = max(len(ref), len(other))
            diffs = []
            for i in range(max_len):
                rv = ref[i] if i < len(ref) else "<missing>"
                ov = other[i] if i < len(other) else "<missing>"
                if rv != ov:
                    diffs.append(f"  Col {i+1}: {ref_name}={rv!r}  {name}={ov!r}")
            mismatches.append(f"\n{ref_name} vs {name}:\n" + "\n".join(diffs))

    if mismatches:
        return False, "Heading mismatches:" + "".join(mismatches)
    return True, ""


# ---------------------------------------------------------------------------
# Consolidation
# ---------------------------------------------------------------------------

def consolidate(found_tabs, wb, num_cols):
    """
    Return a flat list of data rows (no heading) from all present tabs in order.
    Every row is padded/trimmed to exactly num_cols values.
    """
    order = [t for t in REQUIRED_TABS if t in found_tabs]
    result = []
    for tab_name in order:
        data = read_sheet(wb[found_tabs[tab_name]])
        for row in data[1:]:  # skip heading
            padded = (row + [None] * num_cols)[:num_cols]
            if any(v is not None for v in padded):
                result.append(padded)
    return result


# ---------------------------------------------------------------------------
# Column reordering (pure Python list manipulation — no in-place cell shifting)
# ---------------------------------------------------------------------------

def reorder_columns(headings, rows):
    """
    Move Focus Keyword/KW → position 0, Writer → 1, Blog Team → 2.
    Columns not found are skipped. Nothing is deleted.
    Returns (new_headings, new_rows).
    """
    def find(name_set):
        for i, h in enumerate(headings):
            if h and h.strip().lower() in name_set:
                return i
        return None

    fk_idx = find(FOCUS_KW_NAMES)
    wr_idx = find({"writer"})
    bt_idx = find({"blog team"})

    priority = []
    for idx in (fk_idx, wr_idx, bt_idx):
        if idx is not None and idx not in priority:
            priority.append(idx)

    new_order = priority + [i for i in range(len(headings)) if i not in priority]

    new_headings = [headings[i] for i in new_order]
    new_rows = [[row[i] for i in new_order] for row in rows]
    return new_headings, new_rows


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def heading_is_date(heading):
    if not heading:
        return False
    h = heading.strip().lower()
    return any(kw in h for kw in DATE_HEADING_KEYWORDS)


def value_is_date(val):
    return isinstance(val, (datetime, date))


def parse_date(val):
    """Return a datetime if val can be parsed as a date, else None."""
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    if not isinstance(val, str) or not val.strip():
        return None
    for fmt in (
        "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d",
        "%d-%b-%Y", "%b %d, %Y", "%m-%d-%Y", "%d/%m/%Y",
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(val.strip(), fmt)
        except (ValueError, AttributeError):
            continue
    return None


# ---------------------------------------------------------------------------
# Number formatting (step 8)
# ---------------------------------------------------------------------------

def fix_number_formatting(ws, max_row):
    """
    Set purely numeric columns to General format to strip comma formatting.
    Any column that is a date column (by heading name OR by cell value type)
    is unconditionally skipped.
    """
    for c in range(1, ws.max_column + 1):
        heading = ws.cell(row=1, column=c).value

        if heading_is_date(heading):
            continue

        has_numbers = False
        is_date_col = False

        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=c)
            if cell.value is None:
                continue
            if value_is_date(cell.value):
                is_date_col = True
                break
            if isinstance(cell.value, (int, float)):
                has_numbers = True

        if is_date_col:
            continue

        if has_numbers:
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=c)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "General"


# ---------------------------------------------------------------------------
# Date standardization (step 9)
# ---------------------------------------------------------------------------

def standardize_date_columns(ws, max_row):
    """
    Convert date columns to MM/DD/YYYY text. Skips Assignment Sprint column.
    A column is treated as a date column if ≥50% of sampled non-None values
    are date objects or can be parsed as dates.
    """
    for c in range(1, ws.max_column + 1):
        heading = ws.cell(row=1, column=c).value
        if heading and heading.strip().lower() == ASSIGNMENT_SPRINT:
            continue

        date_count = 0
        sample_count = 0
        for row in range(2, min(max_row + 1, 52)):
            val = ws.cell(row=row, column=c).value
            if val is None:
                continue
            sample_count += 1
            if value_is_date(val) or parse_date(val) is not None:
                date_count += 1

        if sample_count == 0 or date_count / sample_count < 0.5:
            continue

        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=c)
            if cell.value is None:
                continue
            parsed = parse_date(cell.value)
            if parsed:
                cell.value = parsed.strftime("%m/%d/%Y")
                cell.number_format = "@"


# ---------------------------------------------------------------------------
# Assignment Sprint → plain text (step 10)
# ---------------------------------------------------------------------------

def convert_assignment_sprint(ws, max_row):
    """Force every value in the Assignment Sprint column to a plain text string."""
    col_idx = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h and h.strip().lower() == ASSIGNMENT_SPRINT:
            col_idx = c
            break

    if col_idx is None:
        return

    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is None:
            continue
        if isinstance(cell.value, datetime):
            cell.value = cell.value.strftime("%m/%d/%Y")
        elif isinstance(cell.value, date):
            cell.value = datetime(cell.value.year, cell.value.month, cell.value.day).strftime("%m/%d/%Y")
        else:
            cell.value = str(cell.value)
        cell.number_format = "@"


# ---------------------------------------------------------------------------
# Active row validation
# ---------------------------------------------------------------------------

def validate_active_rows(headings, rows):
    """
    An active row is any row that has a value in the Focus Keyword column.
    Returns a list of dicts { 'row': 1-indexed row num, 'missing': [col names] }
    for every active row that is missing a value in any REQUIRED_ACTIVE_COLUMNS.
    Columns not present in the file are silently skipped.
    """
    lower_headings = [h.strip().lower() if h else "" for h in headings]

    # Find Focus KW column
    fk_idx = next(
        (i for i, h in enumerate(lower_headings) if h in FOCUS_KW_NAMES), None
    )
    if fk_idx is None:
        return []

    # Map each required column name to its index (skip if not in file)
    req_cols = [
        (headings[lower_headings.index(col)], lower_headings.index(col))
        for col in REQUIRED_ACTIVE_COLUMNS
        if col in lower_headings
    ]

    issues = []
    for row_idx, row in enumerate(rows):
        fk_val = row[fk_idx] if fk_idx < len(row) else None
        if not fk_val or (isinstance(fk_val, str) and not fk_val.strip()):
            continue  # Not an active row

        missing = [
            label
            for label, col_idx in req_cols
            if col_idx >= len(row)
            or row[col_idx] is None
            or (isinstance(row[col_idx], str) and not row[col_idx].strip())
        ]
        if missing:
            issues.append({"row": row_idx + 1, "missing": missing})

    return issues


# ---------------------------------------------------------------------------
# Additional checks and cleanup
# ---------------------------------------------------------------------------

def strip_whitespace(rows):
    """Replace whitespace-only string values with None in place."""
    for row in rows:
        for i, val in enumerate(row):
            if isinstance(val, str) and not val.strip():
                row[i] = None


def check_empty_tabs(wb, found_tabs):
    """Return warning strings for tabs that exist but have no data rows."""
    warnings = []
    for canonical, actual in found_tabs.items():
        data = read_sheet(wb[actual])
        if len(data) <= 1:
            warnings.append(f"{canonical} tab has no data rows")
    return warnings


def check_duplicate_focus_keywords(headings, rows):
    """Return warning strings for duplicate Focus Keyword values."""
    from collections import defaultdict
    lower_headings = [h.strip().lower() if h else "" for h in headings]
    fk_idx = next((i for i, h in enumerate(lower_headings) if h in FOCUS_KW_NAMES), None)
    if fk_idx is None:
        return []

    occurrences = defaultdict(list)
    for row_idx, row in enumerate(rows):
        val = row[fk_idx] if fk_idx < len(row) else None
        if not val or (isinstance(val, str) and not val.strip()):
            continue
        key = str(val).strip().lower()
        occurrences[key].append((row_idx + 1, str(val).strip()))

    warnings = []
    for entries in occurrences.values():
        if len(entries) > 1:
            display = entries[0][1]
            row_nums = [str(e[0]) for e in entries]
            if len(entries) == 2:
                warnings.append(
                    f'Duplicate Focus Keyword: "{display}" appears in rows {row_nums[0]} and {row_nums[1]}'
                )
            else:
                warnings.append(
                    f'Duplicate Focus Keyword: "{display}" appears {len(entries)} times '
                    f'(rows {", ".join(row_nums)})'
                )
    return warnings


def check_duplicate_rows(rows):
    """Return warning strings for exact duplicate rows."""
    warnings = []
    seen = {}
    for row_idx, row in enumerate(rows):
        key = tuple("" if v is None else str(v) for v in row)
        if all(v == "" for v in key):
            continue  # skip all-empty rows
        if key in seen:
            warnings.append(f"Row {row_idx + 1} is an exact duplicate of row {seen[key]}")
        else:
            seen[key] = row_idx + 1
    return warnings


# ---------------------------------------------------------------------------
# Main processing — raises exceptions instead of calling sys.exit()
# Returns (output_xlsx_path, output_csv_path, warnings) on success
# ---------------------------------------------------------------------------

def process(input_path, skip_missing=False):
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"File not found: {input_path}")

    wb = load_workbook(input_path)
    warnings = []

    # Warn if Import-Ready tab already exists (it will be replaced)
    if any(s.lower() == "import-ready" for s in wb.sheetnames):
        warnings.append("Import-Ready tab already existed in the source file and was replaced")

    # Step 1: Validate tabs
    found_tabs, missing_tabs = find_tabs(wb)

    if missing_tabs:
        if not skip_missing:
            raise MissingTabsError(
                f"Missing tabs: {', '.join(missing_tabs)}\n"
                f"Found: {', '.join(found_tabs.keys())}"
            )

    if not found_tabs:
        raise MissingTabsError(
            "None of the required tabs (Marketing, Sales, Service, Website) were found."
        )

    # Warn about tabs that exist but have no data rows
    warnings.extend(check_empty_tabs(wb, found_tabs))

    # Step 3: Validate headings
    ok, detail = validate_headings(found_tabs, wb)
    if not ok:
        raise HeadingMismatchError(detail)

    # Step 4: Get headings from first present tab
    order = [t for t in REQUIRED_TABS if t in found_tabs]
    first_data = read_sheet(wb[found_tabs[order[0]]])
    headings = first_data[0] if first_data else []
    num_cols = len(headings)

    # Check that every required column exists as a heading
    lower_headings = [h.strip().lower() if h else "" for h in headings]
    absent_cols = [col for col in REQUIRED_ACTIVE_COLUMNS if col not in lower_headings]
    if absent_cols:
        raise MissingColumnsError(
            f"Required column(s) not found in file: {', '.join(absent_cols)}"
        )

    # Step 5: Consolidate data rows
    rows = consolidate(found_tabs, wb, num_cols)

    # Strip whitespace-only cells (silent cleanup)
    strip_whitespace(rows)

    # Step 6: Reorder columns (pure list manipulation)
    headings, rows = reorder_columns(headings, rows)

    # Step 7: Drop columns with empty (None) headings
    nonempty_cols = [i for i, h in enumerate(headings) if h is not None]
    headings = [headings[i] for i in nonempty_cols]
    rows = [[row[i] for i in nonempty_cols] for row in rows]

    # Check for duplicate Focus Keywords
    warnings.extend(check_duplicate_focus_keywords(headings, rows))

    # Check for exact duplicate rows
    warnings.extend(check_duplicate_rows(rows))

    # Check active rows for missing required fields
    for issue in validate_active_rows(headings, rows):
        warnings.append(
            f"Row {issue['row']} — missing required fields: {', '.join(issue['missing'])}"
        )

    # Create / replace Import-Ready tab as first sheet
    if "Import-Ready" in wb.sheetnames:
        del wb["Import-Ready"]
    ir_ws = wb.create_sheet("Import-Ready", 0)

    # Write headings
    for c, val in enumerate(headings, 1):
        ir_ws.cell(row=1, column=c).value = val

    # Write data rows
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ir_ws.cell(row=r, column=c).value = val

    max_row = len(rows) + 1

    # Step 8: Fix number formatting (date columns are never touched)
    fix_number_formatting(ir_ws, max_row)

    # Step 9: Standardize date columns to MM/DD/YYYY text
    standardize_date_columns(ir_ws, max_row)

    # Step 10: Convert Assignment Sprint to plain text
    convert_assignment_sprint(ir_ws, max_row)

    # Step 11: Save outputs
    base, ext = os.path.splitext(input_path)
    output_xlsx = f"{base}_import-ready{ext}"
    output_csv = f"{base}_import-ready.csv"

    wb.save(output_xlsx)

    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ir_ws.iter_rows(
            min_row=1, max_row=ir_ws.max_row, max_col=ir_ws.max_column, values_only=True
        ):
            if all(v is None for v in row):
                continue
            writer.writerow(["" if v is None else v for v in row])

    return output_xlsx, output_csv, warnings


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="BCAP Import Helper — consolidate XLSX tabs into Import-Ready CSV"
    )
    parser.add_argument("input_file", help="Path to the input XLSX file")
    parser.add_argument(
        "--skip-missing-tabs",
        action="store_true",
        help="Proceed even if some required tabs are missing",
    )
    args = parser.parse_args()

    try:
        xlsx, csv_path, issues = process(args.input_file, skip_missing=args.skip_missing_tabs)
        print("SUCCESS")
        print(f"XLSX: {xlsx}")
        print(f"CSV:  {csv_path}")
        if warnings:
            print(f"\n{len(warnings)} warning(s):")
            for w in warnings:
                print(f"  {w}")
    except MissingTabsError as e:
        print(str(e), file=sys.stderr)
        sys.exit(10)
    except HeadingMismatchError as e:
        print(str(e), file=sys.stderr)
        sys.exit(20)
    except MissingColumnsError as e:
        print(str(e), file=sys.stderr)
        sys.exit(30)
    except Exception as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
